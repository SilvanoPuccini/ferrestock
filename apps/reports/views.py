import csv
import io

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db import models
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from django.views.generic import TemplateView

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from apps.core.mixins import AppPermissionMixin
from apps.inventory.models import Product, Category
from apps.movements.models import StockMovement
from apps.suppliers.models import Supplier, PurchaseOrder


def get_filtered_products(request):
    queryset = Product.objects.select_related("category", "supplier").order_by("name")

    q = request.GET.get("q", "").strip()
    category_id = request.GET.get("category", "").strip()
    supplier_id = request.GET.get("supplier", "").strip()
    low_stock = request.GET.get("low_stock", "").strip()
    is_active = request.GET.get("is_active", "").strip()

    if q:
        queryset = queryset.filter(
            models.Q(name__icontains=q)
            | models.Q(code__icontains=q)
            | models.Q(barcode__icontains=q)
        )

    if category_id:
        queryset = queryset.filter(category_id=category_id)

    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    if low_stock:
        queryset = queryset.filter(stock_current__lte=models.F("stock_minimum"))

    if is_active == "1":
        queryset = queryset.filter(is_active=True)
    elif is_active == "0":
        queryset = queryset.filter(is_active=False)

    return queryset


def get_filtered_movements(request):
    queryset = StockMovement.objects.select_related("product", "user").order_by("-created_at")

    q = request.GET.get("q", "").strip()
    movement_type = request.GET.get("movement_type", "").strip()
    product_id = request.GET.get("product", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q:
        queryset = queryset.filter(
            models.Q(product__name__icontains=q)
            | models.Q(product__code__icontains=q)
            | models.Q(reason__icontains=q)
            | models.Q(reference__icontains=q)
        )

    if movement_type:
        queryset = queryset.filter(movement_type=movement_type)

    if product_id:
        queryset = queryset.filter(product_id=product_id)

    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)

    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    return queryset


def get_filtered_purchase_orders(request):
    queryset = PurchaseOrder.objects.select_related("supplier", "created_by").prefetch_related("items").order_by("-created_at")

    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    supplier_id = request.GET.get("supplier", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q:
        queryset = queryset.filter(
            models.Q(number__icontains=q)
            | models.Q(supplier__name__icontains=q)
            | models.Q(notes__icontains=q)
        )

    if status:
        queryset = queryset.filter(status=status)

    if supplier_id:
        queryset = queryset.filter(supplier_id=supplier_id)

    if date_from:
        queryset = queryset.filter(created_at__date__gte=date_from)

    if date_to:
        queryset = queryset.filter(created_at__date__lte=date_to)

    return queryset


class StockReportView(AppPermissionMixin, TemplateView):
    permission_required = "inventory.view_product"
    template_name = "reports/stock_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        products = get_filtered_products(self.request)
        export_querystring = self.request.GET.urlencode()

        context["products"] = products
        context["categories"] = Category.objects.filter(is_active=True).order_by("name")
        context["suppliers"] = Supplier.objects.filter(is_active=True).order_by("name")

        context["selected_q"] = self.request.GET.get("q", "")
        context["selected_category"] = self.request.GET.get("category", "")
        context["selected_supplier"] = self.request.GET.get("supplier", "")
        context["selected_low_stock"] = self.request.GET.get("low_stock", "")
        context["selected_is_active"] = self.request.GET.get("is_active", "")

        context["total_products"] = products.count()
        context["total_low_stock"] = products.filter(stock_current__lte=models.F("stock_minimum")).count()
        context["export_querystring"] = export_querystring

        return context


class MovementReportView(AppPermissionMixin, TemplateView):
    permission_required = "movements.view_stockmovement"
    template_name = "reports/movement_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        movements = get_filtered_movements(self.request)
        export_querystring = self.request.GET.urlencode()

        context["movements"] = movements
        context["products"] = Product.objects.order_by("name")
        context["movement_types"] = StockMovement.MOVEMENT_TYPES

        context["selected_q"] = self.request.GET.get("q", "")
        context["selected_movement_type"] = self.request.GET.get("movement_type", "")
        context["selected_product"] = self.request.GET.get("product", "")
        context["selected_date_from"] = self.request.GET.get("date_from", "")
        context["selected_date_to"] = self.request.GET.get("date_to", "")
        context["export_querystring"] = export_querystring

        context["total_movements"] = movements.count()
        context["total_entries"] = movements.filter(movement_type=StockMovement.ENTRY).count()
        context["total_exits"] = movements.filter(movement_type=StockMovement.EXIT).count()
        context["total_adjustments"] = movements.filter(movement_type=StockMovement.ADJUSTMENT).count()

        return context


class PurchaseOrderReportView(AppPermissionMixin, TemplateView):
    permission_required = "suppliers.view_purchaseorder"
    template_name = "reports/purchase_report.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        purchase_orders = get_filtered_purchase_orders(self.request)
        export_querystring = self.request.GET.urlencode()

        context["purchase_orders"] = purchase_orders
        context["suppliers"] = Supplier.objects.filter(is_active=True).order_by("name")
        context["status_choices"] = PurchaseOrder.STATUS_CHOICES

        context["selected_q"] = self.request.GET.get("q", "")
        context["selected_status"] = self.request.GET.get("status", "")
        context["selected_supplier"] = self.request.GET.get("supplier", "")
        context["selected_date_from"] = self.request.GET.get("date_from", "")
        context["selected_date_to"] = self.request.GET.get("date_to", "")
        context["export_querystring"] = export_querystring

        context["total_orders"] = purchase_orders.count()
        context["total_draft"] = purchase_orders.filter(status=PurchaseOrder.DRAFT).count()
        context["total_sent"] = purchase_orders.filter(status=PurchaseOrder.SENT).count()
        context["total_received"] = purchase_orders.filter(status=PurchaseOrder.RECEIVED).count()
        context["total_cancelled"] = purchase_orders.filter(status=PurchaseOrder.CANCELLED).count()
        context["total_amount"] = sum(order.total_amount for order in purchase_orders)

        return context


@login_required
def export_stock_csv(request):
    if not request.user.has_perm("inventory.view_product"):
        raise PermissionDenied

    products = get_filtered_products(request)

    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="ferrestock_stock_report.csv"'},
    )

    writer = csv.writer(response)
    writer.writerow([
        "Codigo",
        "Nombre",
        "Categoria",
        "Proveedor",
        "Stock actual",
        "Stock minimo",
        "Unidad",
        "Estado",
        "Stock bajo",
    ])

    for product in products:
        writer.writerow([
            product.code,
            product.name,
            product.category.name,
            product.supplier.name,
            product.stock_current,
            product.stock_minimum,
            product.get_unit_measure_display(),
            "Activo" if product.is_active else "Inactivo",
            "Si" if product.is_low_stock else "No",
        ])

    return response


@login_required
def export_stock_xlsx(request):
    if not request.user.has_perm("inventory.view_product"):
        raise PermissionDenied

    products = get_filtered_products(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    headers = [
        "Código",
        "Nombre",
        "Categoría",
        "Proveedor",
        "Stock actual",
        "Stock mínimo",
        "Unidad",
        "Estado",
        "Stock bajo",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for product in products:
        ws.append([
            product.code,
            product.name,
            product.category.name,
            product.supplier.name,
            product.stock_current,
            product.stock_minimum,
            product.get_unit_measure_display(),
            "Activo" if product.is_active else "Inactivo",
            "Sí" if product.is_low_stock else "No",
        ])

    widths = {
        "A": 18,
        "B": 28,
        "C": 20,
        "D": 24,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 12,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename="ferrestock_stock_report.xlsx",
    )


@login_required
def export_stock_pdf(request):
    if not request.user.has_perm("inventory.view_product"):
        raise PermissionDenied

    products = get_filtered_products(request)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("FerreStock - Reporte de Stock", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    data = [[
        "Código",
        "Nombre",
        "Categoría",
        "Proveedor",
        "Stock",
        "Mínimo",
        "Unidad",
        "Estado",
        "Bajo",
    ]]

    for product in products:
        data.append([
            product.code,
            product.name,
            product.category.name,
            product.supplier.name,
            str(product.stock_current),
            str(product.stock_minimum),
            product.get_unit_measure_display(),
            "Activo" if product.is_active else "Inactivo",
            "Sí" if product.is_low_stock else "No",
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[22 * mm, 50 * mm, 32 * mm, 40 * mm, 18 * mm, 18 * mm, 22 * mm, 20 * mm, 16 * mm],
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename="ferrestock_stock_report.pdf",
    )


@login_required
def export_movements_csv(request):
    if not request.user.has_perm("movements.view_stockmovement"):
        raise PermissionDenied

    movements = get_filtered_movements(request)

    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="ferrestock_movement_report.csv"'},
    )

    writer = csv.writer(response)
    writer.writerow([
        "Fecha",
        "Producto",
        "Codigo",
        "Tipo",
        "Cantidad",
        "Motivo",
        "Referencia",
        "Usuario",
    ])

    for movement in movements:
        writer.writerow([
            timezone.localtime(movement.created_at).strftime("%d/%m/%Y %H:%M"),
            movement.product.name,
            movement.product.code,
            movement.get_movement_type_display(),
            movement.quantity,
            movement.reason,
            movement.reference,
            movement.user.username,
        ])

    return response


@login_required
def export_movements_xlsx(request):
    if not request.user.has_perm("movements.view_stockmovement"):
        raise PermissionDenied

    movements = get_filtered_movements(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    headers = [
        "Fecha",
        "Producto",
        "Código",
        "Tipo",
        "Cantidad",
        "Motivo",
        "Referencia",
        "Usuario",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for movement in movements:
        ws.append([
            timezone.localtime(movement.created_at).strftime("%d/%m/%Y %H:%M"),
            movement.product.name,
            movement.product.code,
            movement.get_movement_type_display(),
            movement.quantity,
            movement.reason,
            movement.reference,
            movement.user.username,
        ])

    widths = {
        "A": 20,
        "B": 28,
        "C": 18,
        "D": 14,
        "E": 12,
        "F": 32,
        "G": 18,
        "H": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename="ferrestock_movement_report.xlsx",
    )


@login_required
def export_movements_pdf(request):
    if not request.user.has_perm("movements.view_stockmovement"):
        raise PermissionDenied

    movements = get_filtered_movements(request)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("FerreStock - Reporte de Movimientos", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    data = [[
        "Fecha",
        "Producto",
        "Código",
        "Tipo",
        "Cant.",
        "Motivo",
        "Ref.",
        "Usuario",
    ]]

    for movement in movements:
        data.append([
            timezone.localtime(movement.created_at).strftime("%d/%m/%Y %H:%M"),
            movement.product.name,
            movement.product.code,
            movement.get_movement_type_display(),
            str(movement.quantity),
            movement.reason,
            movement.reference or "-",
            movement.user.username,
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[28 * mm, 42 * mm, 22 * mm, 18 * mm, 14 * mm, 55 * mm, 20 * mm, 22 * mm],
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename="ferrestock_movement_report.pdf",
    )


@login_required
def export_purchase_orders_csv(request):
    if not request.user.has_perm("suppliers.view_purchaseorder"):
        raise PermissionDenied

    purchase_orders = get_filtered_purchase_orders(request)

    response = HttpResponse(
        content_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="ferrestock_purchase_report.csv"'},
    )

    writer = csv.writer(response)
    writer.writerow([
        "Numero",
        "Proveedor",
        "Estado",
        "Fecha estimada",
        "Creada por",
        "Creada el",
        "Total",
    ])

    for order in purchase_orders:
        writer.writerow([
            order.number,
            order.supplier.name,
            order.get_status_display(),
            order.expected_date or "",
            order.created_by.username,
            timezone.localtime(order.created_at).strftime("%d/%m/%Y %H:%M"),
            order.total_amount,
        ])

    return response


@login_required
def export_purchase_orders_xlsx(request):
    if not request.user.has_perm("suppliers.view_purchaseorder"):
        raise PermissionDenied

    purchase_orders = get_filtered_purchase_orders(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    headers = [
        "Número",
        "Proveedor",
        "Estado",
        "Fecha estimada",
        "Creada por",
        "Creada el",
        "Total",
    ]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for order in purchase_orders:
        ws.append([
            order.number,
            order.supplier.name,
            order.get_status_display(),
            str(order.expected_date or ""),
            order.created_by.username,
            timezone.localtime(order.created_at).strftime("%d/%m/%Y %H:%M"),
            float(order.total_amount),
        ])

    widths = {
        "A": 18,
        "B": 26,
        "C": 16,
        "D": 16,
        "E": 18,
        "F": 20,
        "G": 14,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename="ferrestock_purchase_report.xlsx",
    )


@login_required
def export_purchase_orders_pdf(request):
    if not request.user.has_perm("suppliers.view_purchaseorder"):
        raise PermissionDenied

    purchase_orders = get_filtered_purchase_orders(request)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("FerreStock - Reporte de Compras", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(
            f"Generado: {timezone.localtime().strftime('%d/%m/%Y %H:%M')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    data = [[
        "Número",
        "Proveedor",
        "Estado",
        "Fecha estimada",
        "Creada por",
        "Creada el",
        "Total",
    ]]

    for order in purchase_orders:
        data.append([
            order.number,
            order.supplier.name,
            order.get_status_display(),
            str(order.expected_date or "-"),
            order.created_by.username,
            timezone.localtime(order.created_at).strftime("%d/%m/%Y %H:%M"),
            str(order.total_amount),
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[24 * mm, 50 * mm, 22 * mm, 24 * mm, 28 * mm, 32 * mm, 18 * mm],
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return FileResponse(
        buffer,
        as_attachment=True,
        filename="ferrestock_purchase_report.pdf",
    )
